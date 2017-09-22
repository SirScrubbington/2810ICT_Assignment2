# Filename: numpy_temp.py
# Author: Damon Murdoch
# Date: 22/09/2017
# Purpose: Generate a new worksheet 'Comparison' (if it doesn't already exist) in 'World Temperature.xlsx' if it exists and inserts Australian average temperature data for the national
# average and for each individual state. Generates a graph to display the data using MatPlotLib.

from sql_temp import *
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import os.path

if __name__ == '__main__':

    newsheet = False
    if os.path.isfile("World Temperature.xlsx"):
        wb = load_workbook("World Temperature.xlsx")
        # If 'comparison' doesn't exist, make a new sheet
        if 'Comparison' not in wb.sheetnames:
            sh = wb.create_sheet("Comparison");
            sh['B1'] = 'Year'
            sh['C1'] = 'Average Temperature'
            sh['A1'] = 'State'
            newsheet = True

        # Connect to database and select average yearly temperatures from Australian stated and the national average
        conn = create_db("data.db")

        sql = "SELECT  STATE, strftime('%Y',_DATE), AVG(AVGTEMP) FROM STATE WHERE COUNTRY ='Australia' GROUP BY STATE, strftime('%Y',_DATE) "
        r = select_from_database(conn, sql)

        sql2 = "SELECT COUNTRY,strftime('%Y',_DATE),AVG(AVGTEMP) FROM COUNTRY WHERE COUNTRY = 'Australia' GROUP BY COUNTRY,strftime('%Y',_DATE)"
        r2 = select_from_database(conn,sql2)

        # Append all of the items from r into r2, and then set r to equal r2
        for i in r:
            r2.append(i)
        r = r2

        # d contains [[year data list],[temperature data list]]
        d = {}

        temp = 0

        # For every item in r
        for i in range(0, len(r)):
            if int(r[i][1]) > 0:
                # If 'Comparison' is new, insert data into worksheet
                if newsheet:
                    sh['A' + str(i + 2)] = str(r[i][0])
                    sh['B' + str(i + 2)] = r[i][1]
                    sh['C' + str(i + 2)] = r[i][2]

                # If new index, declare the arraylist
                if r[i][0] not in d.keys():
                    d[r[i][0]] = [[], []]

                # append year data and temp data to the data list
                if (r[i][2] != 0):
                    d[r[i][0]][0].append(r[i][1])
                    d[r[i][0]][1].append(r[i][2])
            pass


        wb.save("World Temperature.xlsx")

        # Generate Matplot
        for k, v in d.items():
            plt.plot(v[0], v[1], label=k)

        plt.title("Average Temperature in Australian States Against National Average")
        plt.xlabel('Year')
        plt.ylabel('Average Temperature (Celsius)')
        plt.legend()
        plt.show()

    else:
        print("File not found: World Temperature.xlsx")