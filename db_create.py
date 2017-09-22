# Filename: db_create.py
# Author: Damon Murdoch
# Date: 22/09/2017
# Purpose: Generates a database from worksheets in (working dir)/xl
# Notes: Typically takes 5-10 minutes to generate a fresh database.

import sqlite3
from openpyxl import *
from sqlite3 import Error

# create_db(db_file): conn
# Connects to a database with nane 'db_file', or
# creates a new database if it does not exist.
def create_db(db_file):
    try:
        conn = sqlite3.connect(db_file)
        return conn

    except Error as e:
        print(e)

# execite_sql(conn,sql): void
# executes a given sql satement
# on connection 'conn'.
def execute_sql(conn,sql):
    try:
        c = conn.cursor()
        c.execute(sql)
    except Error as e:
        print(e)

if __name__ == '__main__':

    # Global Land Temperatures by City
    create_table_gltbmc = "CREATE TABLE CITY(" \
                          "_DATE DATE NOT NULL," \
                          "AVGTEMP DOUBLE NOT NULL," \
                          "UNCERT DOUBLE NOT NULL," \
                          "CITY VARCHAR(16) NOT NULL," \
                          "COUNTRY VARCHAR(16)," \
                          "LATITUDE VARCHAR(8) NOT NULL," \
                          "LONGITUDE VARCHAR(8) NOT NULL," \
                          "PRIMARY KEY(_DATE,CITY));"

    # Global Land Temperatures By State
    create_table_gltbs = "CREATE TABLE STATE(" \
                         "_DATE DATE NOT NULL," \
                         "AVGTEMP DOUBLE NOT NULL," \
                         "UNCERT DOUBLE NOT NULL," \
                         "STATE VARCHAR(16) NOT NULL," \
                         "COUNTRY VARCHAR(16)," \
                         "PRIMARY KEY(_DATE,STATE));"

    # Global Land Temperatures by Country
    create_table_gltbc = "CREATE TABLE COUNTRY(" \
                          "_DATE DATE NOT NULL," \
                          "AVGTEMP DOUBLE NOT NULL," \
                          "UNCERT DOUBLE NOT NULL," \
                          "COUNTRY VARCHAR(16) NOT NULL," \
                          "PRIMARY KEY(_DATE,COUNTRY));"

    conn = create_db("data.db")

    # If connction was created successfully
    if conn is not None:

        # Create Tables
        execute_sql(conn,create_table_gltbmc)
        execute_sql(conn,create_table_gltbs)
        execute_sql(conn,create_table_gltbc)

        # Load data from Workbooks

        # Cities
        print("loading cities workbook...")
        data_gltbmc = load_workbook("xl/GlobalLandTemperaturesByMajorCity.xlsx")
        data_gltbmc_ws = data_gltbmc.active
        print("done.")

        # Stat es
        print("loading states workbook...")
        data_gltbs = load_workbook("xl/GlobalLandTemperaturesByState.xlsx")
        data_gltbs_ws = data_gltbs.active
        print("done.")

        # Countries
        print("loading countries workbook...")
        data_gltbc = load_workbook("xl/GlobalLandTemperaturesByCountry.xlsx")
        data_gltbc_ws = data_gltbc.active
        print("done.")

        # Load Workbook data into tables

        # Cities
        print("reading cities into database...")
        for i in data_gltbmc_ws:
            sql = """INSERT INTO CITY (_DATE,AVGTEMP,UNCERT,CITY,COUNTRY,LATITUDE,LONGITUDE) VALUES ("{vDATE}","{vAVGTEMP}","{vUNCERT}","{vCITY}","{vCOUNTRY}","{vLATITUDE}","{vLONGITUDE}");"""
            sql = sql.format(vDATE=i[0].value,vAVGTEMP=i[1].value,vUNCERT=i[2].value,vCITY=i[3].value,vCOUNTRY=i[4].value,vLATITUDE=i[5].value,vLONGITUDE=i[6].value)
            #print(sql)
            execute_sql(conn,sql)
            pass
        print("done.")

        # States
        print("reading states into database...")
        for i in data_gltbs_ws:
            sql = """INSERT INTO STATE(_DATE,AVGTEMP,UNCERT,STATE,COUNTRY) VALUES ("{vDATE}","{vAVGTEMP}","{vUNCERT}","{vSTATE}","{vCOUNTRY}");"""
            sql = sql.format(vDATE=i[0].value,vAVGTEMP=i[1].value,vUNCERT=i[2].value,vSTATE=i[3].value,vCOUNTRY=i[4].value)
            #print(sql)
            execute_sql(conn, sql)
            pass
        print("done")

        # Countries
        for i in data_gltbc_ws:
            sql = """INSERT INTO COUNTRY(_DATE,AVGTEMP,UNCERT,COUNTRY) VALUES ("{vDATE}","{vAVGTEMP}","{vUNCERT}","{vCOUNTRY}");"""
            sql = sql.format(vDATE=i[0].value,vAVGTEMP=i[1].value,vUNCERT=i[2].value,vCOUNTRY=i[3].value)
            #print(sql)
            execute_sql(conn, sql)
            pass
        print("done")

        # Save the data and close the connection
        conn.commit()
        conn.close()