# Filename: excel_temp.py
# Author: Damon Murdoch
# Date: 22/09/2017
# Purpose: Generates new workbook 'World Temperature.xlsx' containing Chinese city temperature data, with a graph in the active worksheet.

from sql_temp import *
import matplotlib.pyplot as plt
from openpyxl import *
from openpyxl.chart import (LineChart,Reference,)

if __name__ == '__main__':

    # Create new workbook in memory
    wb = Workbook()
    wb_name = 'World Temperature.xlsx'
    sh = wb.active
    sh.title="Temperature By City"

    # Connect to database and select data from CITY
    conn = create_db("data.db")
    query = "SELECT  CITY, strftime('%Y',_DATE), AVG(AVGTEMP) FROM CITY WHERE COUNTRY ='China' AND AVGTEMP NOT NULL AND _DATE > '0000-00-00' GROUP BY strftime('%Y',_DATE),CITY ORDER BY strftime('%Y',_DATE) DESC"
    r = select_from_database(conn,query)

    # Sort data by year, then by city (city prioritised over year)
    r = sorted(r, key=lambda x: x[1])
    r = sorted(r, key=lambda x: x[0])

    years = []
    cities = []
    temp = {}

    # Generates a list of unique cities, years, and
    # a dictionary containing the intersections of both
    for i in r:
        if i[0] not in cities:
            cities.append(i[0])
        if i[1] not in years:
            years.append(i[1])
        temp[(i[0],i[1])]=i[2]


    # Insert Cities as Column Names
    i=0
    for col in sh.iter_cols(min_row=1,max_row=1,min_col=2,max_col=len(cities)):
        for cell in col:
            cell.value=cities[i]
            i+=1

    # Insert Years as Row Names
    i=0
    for row in sh.iter_rows(min_row=2,max_row=len(years),min_col=1,max_col=1):
        for cell in row:
            cell.value=years[i]
            i+=1

    # Insert the Table Data
    i=0
    j=0
    for row in sh.iter_rows(min_row=2,max_row=len(years),min_col=2,max_col=len(cities)):
        for cell in row:
            try:
                cell.value=temp[(cities[i],years[j])]
            except KeyError:
                pass
            if cell.value==0:
                cell.value=''
            i+=1
        j+=1
        i=0

    # Generate the Graph
    c = LineChart()
    c.title = "Average Temperature in Major Chinese Cities"
    c.style = 10
    c.y_axis.title = "Average Temperature (Celsius)"
    c.x_axis.title = "Year"
    c.x_axis.auto=True

    # Add data from worksheet to Graph
    r = Reference(sh,min_col=2,min_row=1,max_row=len(years),max_col=len(cities))
    c.add_data(r,titles_from_data=True)

    # Add Graph to Worksheet and save
    sh.add_chart(c,"Q1")
    wb.save(wb_name)
